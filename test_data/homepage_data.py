import openpyxl

class HomePageData:

    registration_data = [{"firstname":"tejaswi","email":"teju","gender":"Female"},
                         {"firstname":"anil","email":"gahan","gender":"Male"}]

    #above is hard coded data set, instead we can get data from excel(data_from_excel.py)

    @staticmethod #object creation is not required to access this method
    def get_excel_data(test_case_name):
        book = openpyxl.load_workbook("c:\\users\\...location of excel file")  # loading local excel
        sheet = book.active  # selecting active sheet in the excel

        data_dict = {}
        # to get all data  in excel
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:  # if we need a particular row data
                for j in range(1, sheet.max_column + 1):
                    print(sheet.cell(row=i, column=j).value)
                    data_dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i,column=j).value  # to create a data dict that can be passed to our test case
        return [data_dict] #returning data in list of dict, as thats the way to pass parameterized fixtures
