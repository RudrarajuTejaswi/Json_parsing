import openpyxl
book = openpyxl.load_workbook("c:\\users\\...location of excel file") #loading local excel
sheet = book.active #selecting active sheet in the excel
cell = sheet.cell(row=1,column=2) #selecting particular cell
print(cell.value)
print(sheet['A5'].value) #another way of getting value

#inputting value to the excel
sheet.cell(row=1,column=2).value = "Tejaswi"
print(sheet.max_row) # max no of rows available with data
print(sheet.max_column)# max no of cols available with data

data_dict ={}
#to get all data  in excel
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "TestCase1":# if we need a particular row data
        for j in range(1,sheet.max_column+1):
            print(sheet.cell(row=i,column=j).value)
            data_dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value #to create a data dict that can be passed to our test case






